import tempfile
import os
from datetime import date
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO

from comics.models import Comic, Review, UserSettings
from comics.forms import ReviewForm
from users.forms import RegisterForm, UserSettingsForm
from admin_export.forms import ExportForm

# ---------------------- Тесты моделей ----------------------
class BaseModelTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='testuser', password='12345')
        self.comic = Comic.objects.create(
            title='Test Comic',
            description='Test description',
            genre='superhero',
            release_date=date(2023, 1, 1)
        )

    def test_comic_creation(self):
        self.assertEqual(self.comic.title, 'Test Comic')
        self.assertIsNotNone(self.comic.created_at)
        self.assertIsNotNone(self.comic.updated_at)
        self.assertTrue(self.comic.created_at <= self.comic.updated_at)

    def test_review_creation(self):
        review = Review.objects.create(
            user=self.user,
            comic=self.comic,
            rating=5,
            comment='Great!'
        )
        self.assertEqual(review.user.username, 'testuser')
        self.assertEqual(review.comic.title, 'Test Comic')
        self.assertEqual(review.rating, 5)

    def test_user_settings_creation(self):
        settings = UserSettings.objects.create(user=self.user, theme='dark', font_size='large')
        self.assertEqual(settings.theme, 'dark')
        self.assertEqual(settings.font_size, 'large')

    def test_unique_review_constraint(self):
        Review.objects.create(user=self.user, comic=self.comic, rating=4)
        with self.assertRaises(Exception):
            Review.objects.create(user=self.user, comic=self.comic, rating=5)
    
    def test_rating_validation(self):
        review = Review(user=self.user, comic=self.comic, rating=6)
        with self.assertRaises(Exception):
            review.full_clean()  # валидация на уровне модели

# ---------------------- Тесты форм ----------------------
class FormTests(TestCase):
    def test_valid_review_form(self):
        form = ReviewForm(data={'rating': 4, 'comment': 'Nice comic'})
        self.assertTrue(form.is_valid())

    def test_invalid_review_form(self):
        form = ReviewForm(data={'rating': 6, 'comment': ''})
        self.assertFalse(form.is_valid())
        self.assertIn('rating', form.errors)

    def test_valid_register_form(self):
        form = RegisterForm(data={
            'username': 'newuser',
            'email': 'new@example.com',
            'password1': 'complexpass123',
            'password2': 'complexpass123'
        })
        self.assertTrue(form.is_valid())

    def test_invalid_register_form(self):
        form = RegisterForm(data={
            'username': 'newuser',
            'email': 'notanemail',
            'password1': '123',
            'password2': '123'
        })
        self.assertFalse(form.is_valid())
        self.assertIn('email', form.errors)

# ---------------------- Тесты представлений и доступов ----------------------
class ViewAccessTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='user', password='userpass')
        self.admin = User.objects.create_superuser(username='admin', password='adminpass', email='admin@ex.com')
        self.comic = Comic.objects.create(
            title='Test',
            description='Desc',
            release_date=date(2023,1,1)
        )

    def test_homepage_status(self):
        response = self.client.get(reverse('comic_list'))
        self.assertEqual(response.status_code, 200)

    def test_comic_detail_status(self):
        response = self.client.get(reverse('comic_detail', args=[self.comic.pk]))
        self.assertEqual(response.status_code, 200)

    def test_guest_cannot_edit_review(self):
        response = self.client.get(reverse('edit_review', args=[1]))
        self.assertEqual(response.status_code, 302)  # redirect to login

    def test_user_can_access_settings(self):
        self.client.login(username='user', password='userpass')
        response = self.client.get(reverse('user_settings'))
        self.assertEqual(response.status_code, 200)

    def test_admin_can_access_export(self):
        self.client.login(username='admin', password='adminpass')
        response = self.client.get(reverse('export_report'))
        self.assertEqual(response.status_code, 200)

    def test_non_admin_cannot_access_export(self):
        self.client.login(username='user', password='userpass')
        response = self.client.get(reverse('export_report'))
        self.assertEqual(response.status_code, 302)  # redirect to login (staff required)

# ---------------------- Интеграционный тест основной услуги (отзывы) ----------------------
class ReviewFlowTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='tester', password='testpass')
        self.comic = Comic.objects.create(
            title='Integration Comic',
            description='Desc',
            release_date=date(2023,1,1)
        )

    def test_full_review_cycle(self):
        # 1. Login
        self.client.login(username='tester', password='testpass')
        # 2. Post a review
        response = self.client.post(reverse('comic_detail', args=[self.comic.pk]), {
            'rating': 5,
            'comment': 'Excellent!'
        })
        self.assertEqual(response.status_code, 302)  # redirect after post
        # 3. Verify review in DB
        review = Review.objects.get(comic=self.comic, user=self.user)
        self.assertEqual(review.rating, 5)
        self.assertEqual(review.comment, 'Excellent!')
        # 4. Edit review
        edit_url = reverse('edit_review', args=[review.pk])
        response = self.client.post(edit_url, {'rating': 4, 'comment': 'Updated'})
        self.assertEqual(response.status_code, 302)
        review.refresh_from_db()
        self.assertEqual(review.rating, 4)
        self.assertEqual(review.comment, 'Updated')
        # 5. Count reviews before and after (should remain 1)
        self.assertEqual(Review.objects.count(), 1)

# ---------------------- Тест дополнительного функционала (поиск и фильтрация) ----------------------
class SearchFilterTest(TestCase):
    def setUp(self):
        Comic.objects.create(title='Superman', genre='superhero', release_date=date(2023,1,1))
        Comic.objects.create(title='Batman', genre='superhero', release_date=date(2023,2,1))
        Comic.objects.create(title='Dragon Ball', genre='manga', release_date=date(2023,3,1))

    def test_search_by_title(self):
        response = self.client.get(reverse('comic_list'), {'q': 'Superman'})
        self.assertContains(response, 'Superman')
        self.assertNotContains(response, 'Batman')

    def test_filter_by_genre(self):
        response = self.client.get(reverse('comic_list'), {'genre': 'manga'})
        self.assertContains(response, 'Dragon Ball')
        self.assertNotContains(response, 'Superman')

    def test_sort_by_title(self):
        response = self.client.get(reverse('comic_list'), {'sort': 'title'})
        content = response.content.decode()
        self.assertTrue(content.index('Batman') < content.index('Dragon Ball') < content.index('Superman'))

# ---------------------- Тест административной выгрузки XLSX ----------------------
class AdminExportTest(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(username='admin', password='adminpass')
        self.client.login(username='admin', password='adminpass')
        self.comic = Comic.objects.create(title='Export Comic', description='Desc', release_date=date(2023,1,1))

    def test_export_xlsx(self):
        # Prepare POST data: select table 'comic' and fields 'title', 'description'
        post_data = {
            'tables': ['comic'],
            'fields_comic': ['title', 'description']
        }
        response = self.client.post(reverse('export_report'), post_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Read the Excel file
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        # Check headers
        self.assertEqual(ws['A1'].value, 'title')
        self.assertEqual(ws['B1'].value, 'description')
        # Check data
        self.assertEqual(ws['A2'].value, 'Export Comic')
        self.assertEqual(ws['B2'].value, 'Desc')
        # Ensure only selected fields are present
        self.assertEqual(ws.max_column, 2)

# ---------------------- Тесты безопасности ----------------------
class SecurityTests(TestCase):
    def setUp(self):
        self.client = Client()
        # Создаём хотя бы один комикс, чтобы таблица не была пустой
        self.comic = Comic.objects.create(
            title='Safe Comic',
            description='Test',
            release_date=date(2023, 1, 1)
        )

    def test_sql_injection_in_search(self):
        initial_count = Comic.objects.count()
        malicious = "'; DROP TABLE comics_comic; --"
        response = self.client.get(reverse('comic_list'), {'q': malicious})
        self.assertEqual(response.status_code, 200)
        # Таблица должна существовать, количество записей не изменилось
        self.assertEqual(Comic.objects.count(), initial_count)

    def test_xss_in_comment(self):
        self.user = User.objects.create_user(username='xss', password='xsspass')
        self.client.login(username='xss', password='xsspass')
        self.comic = Comic.objects.create(title='Test', description='Desc', release_date=date(2023,1,1))
        xss_script = "<script>alert('XSS')</script>"
        response = self.client.post(reverse('comic_detail', args=[self.comic.pk]), {
            'rating': 3,
            'comment': xss_script
        })
        self.assertEqual(response.status_code, 302)
        # Check that the stored comment is escaped when rendered
        review = Review.objects.get(comment=xss_script)
        detail_response = self.client.get(reverse('comic_detail', args=[self.comic.pk]))
        self.assertNotIn('<script>', detail_response.content.decode())
        self.assertIn('&lt;script&gt;', detail_response.content.decode())

    def test_passwords_hashed(self):
        user = User.objects.create_user(username='hashme', password='secret')
        # Password should not be stored in plaintext
        self.assertNotEqual(user.password, 'secret')
        self.assertTrue(user.password.startswith('pbkdf2_'))  # Django default hasher