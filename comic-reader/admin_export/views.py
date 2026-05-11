from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from .forms import ExportForm
from comics.models import Comic, Review, UserSettings
from django.contrib.auth.models import User

@staff_member_required
def export_report(request):
    if request.method == 'POST':
        form = ExportForm(request.POST)
        if form.is_valid():
            wb = Workbook()
            tables = form.cleaned_data['tables']
            for table in tables:
                ws = wb.create_sheet(title=dict(ExportForm.MODEL_CHOICES)[table])
                # Получаем выбранные поля
                fields = form.cleaned_data.get(f'fields_{table}', [])
                if not fields:
                    continue
                # Заголовки
                ws.append(fields)
                ws.row_dimensions[1].font = Font(bold=True)
                # Данные
                if table == 'comic':
                    queryset = Comic.objects.all().values(*fields)
                elif table == 'review':
                    queryset = Review.objects.all().values(*fields)
                elif table == 'user':
                    queryset = User.objects.all().values(*fields)
                elif table == 'usersettings':
                    queryset = UserSettings.objects.all().values(*fields)
                else:
                    continue
                for row in queryset:
                    ws.append([row.get(f, '') for f in fields])
            # Удаляем стандартный пустой лист
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report.xlsx'
            wb.save(response)
            return response
    else:
        form = ExportForm()
    return render(request, 'admin_export/export_report.html', {'form': form})